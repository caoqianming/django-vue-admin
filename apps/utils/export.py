import xlwt
import time
import os
from django.conf import settings
from datetime import datetime
from openpyxl import Workbook, styles
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter, column_index_from_string


def len_byte(value):
    # 获取字符串长度，一个中文的长度为2
    length = len(value)
    utf8_length = len(value.encode('utf-8'))
    length = (utf8_length - length) / 2 + length
    return int(length)


def export_excel(field_data: list, data: list, FileName: str):
    """
    Excel导出
    :param data: 数据源
    :param field_data: 首行数据源（表头）
    :param file_path: 文件保存路径（默认保存在media路径）
    :param FileName: 文件保存名字
    :return:返回文件的下载url完整路径
    """
    wbk = xlwt.Workbook(encoding='utf-8')
    sheet = wbk.add_sheet('Sheet1', cell_overwrite_ok=True)  # 第二参数用于确认同一个cell单元是否可以重设值。
    style = xlwt.XFStyle()  # 赋值style为XFStyle()，初始化样式
    # 设置居中
    wbk.set_colour_RGB(0x23, 0, 60, 139)
    xlwt.add_palette_colour("custom_colour_35", 0x23)
    tab_al = xlwt.Alignment()
    tab_al.horz = 0x02  # 设置水平居中
    tab_al.vert = 0x01  # 设置垂直居中
    # 设置表头单元格背景颜色
    tab_pattern = xlwt.Pattern()  # 创建一个模式
    tab_pattern.pattern = xlwt.Pattern.SOLID_PATTERN  # 设置其模式为实型
    tab_pattern.pattern_fore_colour = 55
    # 设置单元格内字体样式
    tab_fnt = xlwt.Font()  # 创建一个文本格式，包括字体、字号和颜色样式特性
    tab_fnt.height = 200
    default_width = 14
    tab_fnt.name = u'楷体'  # 设置其字体为微软雅黑
    tab_fnt.colour_index = 1  # 设置其字体颜色
    # 设置单元格下框线样式
    tab_borders = xlwt.Borders()
    tab_borders.left = xlwt.Borders.THIN
    tab_borders.right = xlwt.Borders.THIN
    tab_borders.top = xlwt.Borders.THIN
    tab_borders.bottom = xlwt.Borders.THIN
    tab_borders.left_colour = 23
    tab_borders.right_colour = 23
    tab_borders.bottom_colour = 23
    tab_borders.top_colour = 23
    # 把数据写入excel中
    # 所有表格单元格样式
    # 先生成表头
    style.alignment = tab_al  # 设置居中
    style.pattern = tab_pattern  # 设置表头单元格背景颜色
    style.font = tab_fnt  # 设置单元格内字体样式
    style.borders = tab_borders
    for index, ele in enumerate(field_data):
        sheet.write_merge(0, 0, index, index, ele, style)  # (列开始, 列结束, 行开始, 行结束, '数据内容')

    # 确定栏位宽度
    col_width = []
    for index, ele in enumerate(data):
        for inx, values in enumerate(ele):
            if index == 0:
                col_width.append(len_byte(str(values)))
            else:
                if col_width[inx] < len_byte(str(values)):
                    col_width[inx] = len_byte(str(values))
    # 设置栏位宽度，栏位宽度小于10时候采用默认宽度
    for i in range(len(col_width)):
        if col_width[i] > 10:
            width = col_width[i] if col_width[i] < 36 else 36
            sheet.col(i).width = 256 * (width + 6)
        else:
            sheet.col(i).width = 256 * (default_width)

    row = 1
    # 内容背景颜色
    left_pattern = xlwt.Pattern()  # 创建一个模式
    left_pattern.pattern = xlwt.Pattern.SOLID_PATTERN  # 设置其模式为实型
    left_pattern.pattern_fore_colour = 1

    # 设置单元格内字体样式
    left_fnt = xlwt.Font()  # 创建一个文本格式，包括字体、字号和颜色样式特性
    left_fnt.height = 200
    left_fnt.name = u'楷体'  # 设置其字体为微软雅黑
    left_fnt.colour_index = 0  # 设置其字体颜色

    left_style = style
    left_style.pattern = left_pattern
    left_style.font = left_fnt

    for results in data:
        for index, values in enumerate(results):
            sheet.write(row, index, label=values, style=left_style)
        row += 1

    FileNameF = FileName + datetime.now().strftime('%Y%m%d%H%M%S') + '.xls'
    path = '/media/temp/'
    pathRoot = settings.BASE_DIR + path
    if not os.path.exists(pathRoot):
        os.makedirs(pathRoot)

    path_name = os.path.join(pathRoot, FileNameF)
    wbk.save(path_name)
    return path + FileNameF


def export_excel_img(field_data: list, data: list, FileName: str):
    """
    带有image的Excel导出
    :param data: 数据源
    :param field_data: 首行数据源（表头）{'name':'', 'type':''}
    :param img_field_indexs: 图片字段名index列表
    :param file_path: 文件保存路径（默认保存在media路径）
    :param FileName: 文件保存名字
    :return:返回文件的下载url完整路径
    """
    wb = Workbook()
    ws = wb.active
    imgs = []

    for index, value in enumerate(field_data):
        cell = ws.cell(column=index+1, row=1)
        cell.value = value['name']
        cell.font = styles.Font(bold=True)
        letter = get_column_letter(index+1)
        value['letter'] = letter
        ws.column_dimensions[letter].width = 10  # 修改列宽
        if value['type'] == 'img':
            ws.column_dimensions[letter].width = 15  # 修改列宽

    for i1, v1 in enumerate(data):
        for i2, v2 in enumerate(v1):
            cell = ws.cell(column=i2+1, row=i1+2)
            if v2 and field_data[i2]['type'] == 'img':
                ws.row_dimensions[i1+2].height = 70
                try:
                    img = Image(settings.BASE_DIR + v2)
                    img.width, img.height = (90, 90)
                    imgs.append((img, field_data[i2]['letter'] + str(i1+2)))
                except:  # 这里先不做处理
                    pass 
            else:
                cell.value = v2

    for i in imgs:
        ws.add_image(i[0], i[1])

    FileNameF = FileName + datetime.now().strftime('%Y%m%d%H%M%S') + '.xlsx'
    path = '/temp/'
    pathRoot = settings.BASE_DIR + path
    if not os.path.exists(pathRoot):
        os.makedirs(pathRoot)

    path_name = os.path.join(pathRoot, FileNameF)
    wb.save(path_name)
    return path + FileNameF
